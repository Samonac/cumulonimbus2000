import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import sys


def generate_colored_grid(image_path, num_rows, num_cols):
    # Read the image
    image = cv2.imread(image_path)

    # Get the original dimensions of the image
    original_height, original_width, _ = image.shape

    # Calculate the dimensions of each cell
    cell_height = original_height // num_rows
    cell_width = original_width // num_cols

    # Initialize a blank canvas for the colored grid
    colored_grid = np.zeros((num_rows, num_cols, 3), dtype=np.uint8)

    # Initialize an Excel workbook
    wb = Workbook()
    ws = wb.active
    jsonCoord2Colors = {}

    # Iterate through each cell and calculate the average color
    for i in range(num_rows):
        for j in range(num_cols):
            # Define the region of interest (ROI) for the current cell
            y_start = i * cell_height
            y_end = (i + 1) * cell_height
            x_start = j * cell_width
            x_end = (j + 1) * cell_width

            # Extract the ROI from the image
            roi = image[y_start:y_end, x_start:x_end]

            # Calculate the average color of the ROI
            average_color = np.mean(roi, axis=(0, 1)).astype(np.uint8)

            # Assign the average color to the corresponding cell in the grid
            colored_grid[i, j] = average_color

            # Convert RGB values to hexadecimal color code
            hex_color = '%02x%02x%02x' % tuple(average_color)

            # Set the background color of the cell
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell = ws.cell(row=i + 1, column=j + 1)
            coord = '{};{}'.format(i + 1, j + 1)
            jsonCoord2Colors[coord] = [average_color[0], average_color[1], average_color[2]]
            cell.fill = fill

    # Save the Excel file
    excel_file = image_path.split('.')[0] + "_colors.xlsx"
    print(jsonCoord2Colors)
    wb.save(excel_file)

    return colored_grid, jsonCoord2Colors


def main(image_path='D:/Videos/Frames/frame_24.png'):
    # Get input image path and desired number of rows and columns from the user
    if 'png' not in '{}'.format(image_path):
        image_path = 'D:/Videos/Frames/frame_24.png'
    num_rows = 40
    num_cols = 160

    # Generate the colored grid
    colored_grid, jsonCoord2Colors = generate_colored_grid(image_path, num_rows, num_cols)
    return jsonCoord2Colors
    # Display the colored grid
    # cv2.imshow("Colored Grid", colored_grid)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()


if __name__ == "__main__":
    if len(sys.argv) > 1:
        main(sys.argv[1])
    else:
        main(image_path='D:/Videos/Frames/frame_24.png')
        print("Usage: python script.py <image_path>")


def main_old():
    # Get input image path and desired dimensions from the user
    image_path = 'D:/Videos/Frames/frame_24.png'
    # new_width = int(input("Enter the new width for the grid: "))
    new_width = 160
    # new_height = int(input("Enter the new height for the grid: "))
    new_height = 110

    # Generate the colored grid
    colored_grid = generate_colored_grid(image_path, new_width, new_height)

    # Display the colored grid
    cv2.imshow("Colored Grid", colored_grid)
    cv2.waitKey(0)
    cv2.destroyAllWindows()

